from __future__ import annotations

import csv
import io
from datetime import date, datetime

from .base import SourceAdapter, SourceTable


class CsvAdapter(SourceAdapter):
    name = "csv"
    extensions = (".csv", ".txt")

    def read(self, content: bytes, filename: str) -> SourceTable:
        text = _decode(content)
        sample = text[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
            dialect.delimiter = ";" if sample.count(";") > sample.count(",") else ","

        reader = csv.DictReader(io.StringIO(text), dialect=dialect)
        columns = [c.strip() for c in (reader.fieldnames or [])]
        rows = [
            {(k or "").strip(): (v or "").strip() for k, v in row.items() if k is not None}
            for row in reader
        ]
        return SourceTable(columns=columns, rows=rows)


class ExcelAdapter(SourceAdapter):
    name = "excel"
    extensions = (".xlsx", ".xlsm")

    def read(self, content: bytes, filename: str) -> SourceTable:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            return SourceTable(columns=[], rows=[])

        columns = [str(h).strip() if h is not None else f"col{i + 1}" for i, h in enumerate(header)]
        rows: list[dict[str, str]] = []
        for raw in rows_iter:
            if raw is None or all(v is None or str(v).strip() == "" for v in raw):
                continue
            rows.append({columns[i]: _cell(v) for i, v in enumerate(raw) if i < len(columns)})
        wb.close()
        return SourceTable(columns=columns, rows=rows)


def _cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _decode(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace")


ADAPTERS: list[SourceAdapter] = [CsvAdapter(), ExcelAdapter()]


def adapter_for(filename: str) -> SourceAdapter:
    lowered = filename.lower()
    for adapter in ADAPTERS:
        if any(lowered.endswith(ext) for ext in adapter.extensions):
            return adapter
    raise ValueError(
        "Formato non supportato: usare " + ", ".join(sorted({e for a in ADAPTERS for e in a.extensions}))
    )
